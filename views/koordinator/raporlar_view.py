# -*- coding: utf-8 -*-
"""
Raporlar View - Tam özellikli
PDF ve Excel raporlar
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QFrame,
    QComboBox, QFileDialog, QMessageBox, QGroupBox, QScrollArea
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont
from controllers.sinav_controller import SinavController
from models.database import db
from models.sinav_model import SinavModel
from utils.pdf_generator import PDFGenerator
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class RaporCard(QFrame):
    """Rapor kartı"""

    clicked = Signal()

    def __init__(self, title, description, icon, color, parent=None):
        super().__init__(parent)
        self.title_text = title
        self.desc_text = description
        self.icon_text = icon
        self.color_value = color
        self.init_ui()

    def init_ui(self):
        """UI oluştur"""
        self.setCursor(Qt.PointingHandCursor)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(12)

        # İkon
        icon = QLabel(self.icon_text)
        icon.setFont(QFont("Segoe UI", 40))
        icon.setAlignment(Qt.AlignCenter)

        # Başlık
        title = QLabel(self.title_text)
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #111827;")

        # Açıklama
        desc = QLabel(self.desc_text)
        desc.setFont(QFont("Segoe UI", 10))
        desc.setAlignment(Qt.AlignCenter)
        desc.setStyleSheet("color: #6b7280;")
        desc.setWordWrap(True)

        layout.addWidget(icon)
        layout.addWidget(title)
        layout.addWidget(desc)

        self.setStyleSheet(f"""
            QFrame {{
                background: white;
                border: 2px solid #e5e7eb;
                border-radius: 16px;
            }}
            QFrame:hover {{
                border-color: {self.color_value};
                background: #f9fafb;
            }}
        """)

    def mousePressEvent(self, event):
        """Mouse click"""
        self.clicked.emit()
        super().mousePressEvent(event)


class RaporlarView(QWidget):
    """Raporlar ekranı - Tam özellikli"""

    def __init__(self, user_data):
        super().__init__()
        self.user_data = user_data
        self.controller = SinavController()
        self.sinav_model = SinavModel(db)
        self.pdf_generator = PDFGenerator()
        self.init_ui()

    def init_ui(self):
        """UI oluştur"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(20)

        # Header
        header = QFrame()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("📊 Raporlar")
        title.setFont(QFont("Segoe UI", 22, QFont.Bold))
        title.setStyleSheet("color: #111827;")

        header_layout.addWidget(title)
        header_layout.addStretch()

        layout.addWidget(header)

        # Program seçimi
        program_group = QGroupBox("Sınav Programı Seçin")
        program_group.setFont(QFont("Segoe UI", 11, QFont.Bold))
        program_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #e5e7eb;
                border-radius: 12px;
                margin-top: 12px;
                padding: 16px;
                background: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 0 8px;
                color: #374151;
            }
        """)
        program_layout = QHBoxLayout(program_group)

        self.cmb_program = QComboBox()
        self.cmb_program.setMinimumHeight(44)
        self.cmb_program.setStyleSheet("""
            QComboBox {
                padding: 10px;
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                font-size: 14px;
                background: white;
            }
            QComboBox:focus {
                border-color: #10b981;
            }
        """)
        self.load_programs()

        program_layout.addWidget(QLabel("Program:"))
        program_layout.addWidget(self.cmb_program, 1)

        layout.addWidget(program_group)

        # Rapor kartları
        cards_label = QLabel("Rapor Türleri")
        cards_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        cards_label.setStyleSheet("color: #374151; margin-top: 16px;")
        layout.addWidget(cards_label)

        # Grid layout for cards
        cards_container = QWidget()
        cards_layout = QHBoxLayout(cards_container)
        cards_layout.setSpacing(16)

        # Sınav programı raporu
        card1 = RaporCard(
            "Sınav Programı",
            "Tüm sınavların tarih, saat ve derslik bilgilerini içeren PDF raporu",
            "📅",
            "#10b981"
        )
        card1.clicked.connect(self.generate_sinav_programi_pdf)
        cards_layout.addWidget(card1)

        # Oturma planı raporu
        card2 = RaporCard(
            "Oturma Planları",
            "Her sınav için öğrenci oturma düzenini gösteren PDF raporu",
            "📋",
            "#3b82f6"
        )
        card2.clicked.connect(self.generate_oturma_plani_pdf)
        cards_layout.addWidget(card2)

        # Excel raporu
        card3 = RaporCard(
            "Excel Çıktısı",
            "Sınav programının Excel formatında dışa aktarımı",
            "📊",
            "#f59e0b"
        )
        card3.clicked.connect(self.generate_excel_report)
        cards_layout.addWidget(card3)

        # Özet rapor
        card4 = RaporCard(
            "Özet Rapor",
            "Sınav programı özet bilgileri ve istatistikleri",
            "📈",
            "#8b5cf6"
        )
        card4.clicked.connect(self.generate_summary_pdf)
        cards_layout.addWidget(card4)

        layout.addWidget(cards_container)

        layout.addStretch()

    def load_programs(self):
        """Programları yükle"""
        try:
            bolum_id = self.user_data.get('bolum_id', 1)
            programs = self.controller.get_programs_by_bolum(bolum_id)

            self.cmb_program.clear()
            if not programs:
                self.cmb_program.addItem("Henüz program oluşturulmamış", None)
                return

            for program in programs:
                self.cmb_program.addItem(
                    f"{program['program_adi']} ({program['sinav_tipi']})",
                    program['program_id']
                )

        except Exception as e:
            logger.error(f"Program yükleme hatası: {e}")

    def get_selected_program_id(self):
        """Seçili program ID'sini al"""
        return self.cmb_program.currentData()

    def generate_sinav_programi_pdf(self):
        """Sınav programı PDF raporu"""
        program_id = self.get_selected_program_id()
        if not program_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir program seçin!")
            return

        try:
            # Program bilgilerini al
            program = self.controller.get_program_by_id(program_id)
            sinavlar = self.controller.get_sinavlar_by_program(program_id)

            if not sinavlar:
                QMessageBox.information(self, "Bilgi", "Bu programa ait sınav bulunmuyor.")
                return

            # PDF oluştur
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "PDF Kaydet",
                f"Sinav_Programi_{program['program_adi']}.pdf",
                "PDF Files (*.pdf)"
            )

            if filename:
                self.pdf_generator.generate_sinav_programi(program, sinavlar, filename)
                QMessageBox.information(self, "Başarılı", f"Rapor oluşturuldu:\n{filename}")

        except Exception as e:
            logger.error(f"PDF oluşturma hatası: {e}")
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulamadı:\n{str(e)}")

    def generate_oturma_plani_pdf(self):
        """Oturma planı PDF raporu"""
        program_id = self.get_selected_program_id()
        if not program_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir program seçin!")
            return

        try:
            program = self.controller.get_program_by_id(program_id)
            sinavlar = self.controller.get_sinavlar_by_program(program_id)

            if not sinavlar:
                QMessageBox.information(self, "Bilgi", "Bu programa ait sınav bulunmuyor.")
                return

            # PDF oluştur
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "PDF Kaydet",
                f"Oturma_Plani_{program['program_adi']}.pdf",
                "PDF Files (*.pdf)"
            )

            if filename:
                self.pdf_generator.generate_oturma_planlari(program, sinavlar, filename)
                QMessageBox.information(self, "Başarılı", f"Rapor oluşturuldu:\n{filename}")

        except Exception as e:
            logger.error(f"PDF oluşturma hatası: {e}")
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulamadı:\n{str(e)}")

    def generate_excel_report(self):
        """Excel raporu"""
        program_id = self.get_selected_program_id()
        if not program_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir program seçin!")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            program = self.controller.get_program_by_id(program_id)
            sinavlar = self.controller.get_sinavlar_by_program(program_id)

            if not sinavlar:
                QMessageBox.information(self, "Bilgi", "Bu programa ait sınav bulunmuyor.")
                return

            # Excel dosyası oluştur
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Kaydet",
                f"Sinav_Programi_{program['program_adi']}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if filename:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sınav Programı"

                # Başlık
                ws['A1'] = program['program_adi']
                ws['A1'].font = Font(size=16, bold=True)
                ws.merge_cells('A1:F1')

                # Sütun başlıkları
                headers = ['Ders Kodu', 'Ders Adı', 'Tarih', 'Saat', 'Derslik', 'Öğrenci Sayısı']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')

                # Veriler
                for row_idx, sinav in enumerate(sinavlar, start=4):
                    ws.cell(row=row_idx, column=1).value = sinav.get('ders_kodu', 'N/A')
                    ws.cell(row=row_idx, column=2).value = sinav.get('ders_adi', 'N/A')

                    tarih = sinav.get('tarih')
                    if tarih:
                        if isinstance(tarih, str):
                            ws.cell(row=row_idx, column=3).value = tarih
                        else:
                            ws.cell(row=row_idx, column=3).value = tarih.strftime("%d.%m.%Y")

                    bas_saat = sinav.get('baslangic_saati')
                    bit_saat = sinav.get('bitis_saati')
                    if bas_saat and bit_saat:
                        ws.cell(row=row_idx, column=4).value = f"{bas_saat} - {bit_saat}"

                    ws.cell(row=row_idx, column=5).value = sinav.get('derslik_adi', 'Atanmadı')
                    ws.cell(row=row_idx, column=6).value = sinav.get('ogrenci_sayisi', 0)

                # Sütun genişlikleri
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 15

                wb.save(filename)
                QMessageBox.information(self, "Başarılı", f"Excel raporu oluşturuldu:\n{filename}")

        except ImportError:
            QMessageBox.critical(
                self,
                "Hata",
                "openpyxl kütüphanesi bulunamadı.\n\nTerminalden şu komutu çalıştırın:\npip install openpyxl"
            )
        except Exception as e:
            logger.error(f"Excel oluşturma hatası: {e}")
            QMessageBox.critical(self, "Hata", f"Excel raporu oluşturulamadı:\n{str(e)}")

    def generate_summary_pdf(self):
        """Özet rapor PDF"""
        program_id = self.get_selected_program_id()
        if not program_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir program seçin!")
            return

        try:
            program = self.controller.get_program_by_id(program_id)
            sinavlar = self.controller.get_sinavlar_by_program(program_id)

            if not sinavlar:
                QMessageBox.information(self, "Bilgi", "Bu programa ait sınav bulunmuyor.")
                return

            # İstatistikler hesapla
            toplam_sinav = len(sinavlar)
            toplam_ogrenci = sum(s.get('ogrenci_sayisi', 0) for s in sinavlar)

            # Tarihe göre grupla
            tarihler = {}
            for sinav in sinavlar:
                tarih = sinav.get('tarih')
                if tarih:
                    tarih_str = tarih.strftime("%d.%m.%Y") if not isinstance(tarih, str) else tarih
                    tarihler[tarih_str] = tarihler.get(tarih_str, 0) + 1

            # PDF oluştur
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "PDF Kaydet",
                f"Ozet_Rapor_{program['program_adi']}.pdf",
                "PDF Files (*.pdf)"
            )

            if filename:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import cm

                c = canvas.Canvas(filename, pagesize=A4)
                width, height = A4

                # Başlık
                c.setFont("Helvetica-Bold", 18)
                c.drawString(2*cm, height - 3*cm, "SINAV PROGRAMI ÖZET RAPORU")

                # Program bilgileri
                y = height - 5*cm
                c.setFont("Helvetica-Bold", 12)
                c.drawString(2*cm, y, "Program Bilgileri")

                y -= 0.8*cm
                c.setFont("Helvetica", 10)
                c.drawString(2*cm, y, f"Program Adı: {program['program_adi']}")

                y -= 0.6*cm
                c.drawString(2*cm, y, f"Sınav Tipi: {program['sinav_tipi']}")

                y -= 0.6*cm
                bas_tarih = program['baslangic_tarihi']
                bit_tarih = program['bitis_tarihi']
                if not isinstance(bas_tarih, str):
                    bas_tarih = bas_tarih.strftime("%d.%m.%Y")
                if not isinstance(bit_tarih, str):
                    bit_tarih = bit_tarih.strftime("%d.%m.%Y")
                c.drawString(2*cm, y, f"Tarih Aralığı: {bas_tarih} - {bit_tarih}")

                # İstatistikler
                y -= 1.5*cm
                c.setFont("Helvetica-Bold", 12)
                c.drawString(2*cm, y, "İstatistikler")

                y -= 0.8*cm
                c.setFont("Helvetica", 10)
                c.drawString(2*cm, y, f"Toplam Sınav Sayısı: {toplam_sinav}")

                y -= 0.6*cm
                c.drawString(2*cm, y, f"Toplam Öğrenci Sayısı: {toplam_ogrenci}")

                y -= 0.6*cm
                c.drawString(2*cm, y, f"Sınav Günü Sayısı: {len(tarihler)}")

                # Günlük dağılım
                y -= 1.5*cm
                c.setFont("Helvetica-Bold", 12)
                c.drawString(2*cm, y, "Günlük Sınav Dağılımı")

                y -= 0.8*cm
                c.setFont("Helvetica", 10)
                for tarih, sayi in sorted(tarihler.items()):
                    c.drawString(2*cm, y, f"{tarih}: {sayi} sınav")
                    y -= 0.6*cm

                c.save()
                QMessageBox.information(self, "Başarılı", f"Özet rapor oluşturuldu:\n{filename}")

        except Exception as e:
            logger.error(f"PDF oluşturma hatası: {e}")
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulamadı:\n{str(e)}")
